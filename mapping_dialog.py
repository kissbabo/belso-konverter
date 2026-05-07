"""
KK-Industry – BELSŐ Költségvetés Konverter
==========================================
Grafikus alkalmazás: Excel ajánlat → BELSŐ formátum
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from mapping_dialog import show_mapping_dialog, BELSO_ROLES
try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
    HAS_DND = True
except ImportError:
    HAS_DND = False
import threading
import os
import sys
from pathlib import Path
from datetime import datetime

# ── Konverter logika ──────────────────────────────────────────────────────────
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Color
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter, column_index_from_string

# Stílusok
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
BOLD        = Font(bold=True)
CENTER      = Alignment(horizontal='center', vertical='center', wrap_text=True)
WRAP        = Alignment(wrap_text=True, vertical='top')
THIN        = Side(style='thin')
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

BELSO_HEADERS = [
    ('A','No'), ('B','Tétel megnevezése'), ('C','Mennyiség'), ('D','Egység'),
    ('E','Anyag be'), ('F','EURO'), ('G','Besz.'), ('H','Kedv.'),
    ('I','Kedv. Ár'), ('J','Ax kód'), ('K','Ax'), ('L','Anyag egységár'),
    ('M','No.idő'), ('N','Egység munkadíj'), ('O','Anyag összesen'),
    ('P','Munkadíj összesen'), ('Q','Összesen'),
]
YELLOW_HEADER_COLS = {'E','F','G','H','J','M'}

COL_FORMATS = {
    3:  '#,##0.##',
    5:  '#,##0 "Ft"',
    6:  '#,##0.0 "€"',
    8:  '0%',
    9:  '#,##0 "Ft"',
    11: '0%',
    12: '#,##0 "Ft"',
    13: '#,##0.##',
    14: '#,##0 "Ft"',
    15: '#,##0 "Ft"',
    16: '#,##0 "Ft"',
    17: '#,##0 "Ft"',
}

def make_cf_fill(bg_hex):
    fill = PatternFill(fill_type='solid')
    fill.fgColor = Color('00000000')
    fill.bgColor = Color(bg_hex)
    return fill

def add_conditional_formatting(ws, data_last_row):
    EQ_ZERO  = DifferentialStyle(font=Font(color='FF9C0006'), fill=make_cf_fill('FFFFBFBF'))
    LT_ZERO  = DifferentialStyle(font=Font(color='FF9C0006'), fill=make_cf_fill('FF008000'))
    GT_ZERO  = DifferentialStyle(font=Font(color='FF000000'), fill=make_cf_fill('FFBFFFBF'))
    GRN_FILL = DifferentialStyle(font=Font(color='FF000000'), fill=make_cf_fill('FFBFFFBF'))
    for col_letter in ('E', 'M'):
        cf = f'{col_letter}2:{col_letter}{data_last_row}'
        ws.conditional_formatting.add(cf, Rule(type='cellIs', operator='equal',   formula=['0'], dxf=EQ_ZERO,  priority=4))
        ws.conditional_formatting.add(cf, Rule(type='cellIs', operator='lessThan', formula=['0'], dxf=LT_ZERO,  priority=5))
        ws.conditional_formatting.add(cf, Rule(type='cellIs', operator='greaterThan', formula=['0'], dxf=GT_ZERO, priority=6))
    ws.conditional_formatting.add(f'E2:E{data_last_row}',
        Rule(type='expression', formula=['F2>0'], dxf=GRN_FILL, priority=1))

def set_header(ws):
    for col_idx, (col_letter, header) in enumerate(BELSO_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = BOLD; cell.alignment = CENTER; cell.border = BORDER
        if col_letter in YELLOW_HEADER_COLS:
            cell.fill = YELLOW_FILL

def write_data_row(ws, out_row, ssz, leiras, mennyiseg, egyseg):
    r = out_row
    ws.cell(r, 1, value=ssz).border = BORDER
    cell_b = ws.cell(r, 2, value=leiras)
    cell_b.alignment = WRAP; cell_b.border = BORDER
    cell_c = ws.cell(r, 3, value=mennyiseg)
    cell_c.alignment = CENTER; cell_c.border = BORDER
    cell_d = ws.cell(r, 4, value=egyseg)
    cell_d.alignment = CENTER; cell_d.border = BORDER
    ws.cell(r, 5, value=0).border = BORDER
    ws.cell(r, 6, value=None).border = BORDER
    ws.cell(r, 7, value=None).border = BORDER
    ws.cell(r, 8, value=0).border = BORDER
    ws.cell(r, 9,  value=f'=IF(F{r}>0,F{r}*BELSŐ!$F$3,E{r})*(1-H{r})').border = BORDER
    ws.cell(r, 10, value='Anyag').border = BORDER
    ws.cell(r, 11, value=f'=VLOOKUP(J{r},BELSŐ!C:D,2,FALSE)').border = BORDER
    ws.cell(r, 12, value=f'=I{r}*(1+K{r})').border = BORDER
    ws.cell(r, 13, value=0).border = BORDER                          # M: No.ido (input)
    ws.cell(r, 14, value=f'=M{r}*BELSŐ!$D$3').border = BORDER        # N: Egyseg munkadij = No.ido x oradij
    ws.cell(r, 15, value=f'=C{r}*L{r}').border = BORDER              # O: Anyag osszesen = Mennyiseg x Anyag egysar
    ws.cell(r, 16, value=f'=C{r}*N{r}').border = BORDER              # P: Munkadij osszesen = Mennyiseg x Egyseg munkadij
    ws.cell(r, 17, value=f'=O{r}+P{r}').border = BORDER              # Q: Osszesen = Anyag + Munkadij
    for col_idx, fmt in COL_FORMATS.items():
        ws.cell(r, col_idx).number_format = fmt

# Összesítő sorok szűrési kulcsszavai
SKIP_WORDS = ['összesen', 'mindösszesen', 'opció', 'kiegészít', 'subtotal', 'total']

# ── Mapping Excel generálás és beolvasás ──────────────────────────────────────

# Az 5 keresett BELSŐ oszlop szerepe és megjelenítési neve
BELSO_ROLES = [
    ('mennyiseg',  'Mennyiség'),
    ('egyseg',     'Egység'),
    ('anyag_egy',  'Anyag egységár'),
    ('dij_egy',    'Díj egységár'),
    ('anyag_ossz', 'Anyag összár'),
    ('dij_ossz',   'Díj összár'),
]

def detect_full_mapping(ws):
    """
    Kibővített oszlop mapping: megkeresi az 5 BELSŐ szerephez tartozó
    forrás oszlopot (cellacímet pl. C1, D1) és a fejléc sor számát.
    """
    header_row = detect_header_row(ws)
    base = detect_column_mapping(ws, header_row)

    # Alap szerepek
    result = {
        'header_row': header_row,
        'ssz':       base.get('ssz'),
        'leiras':    base.get('leiras'),
        'mennyiseg': base.get('mennyiseg'),
        'egyseg':    base.get('egyseg'),
        'anyag_egy': None,
        'dij_egy':   None,
        'anyag_ossz':None,
        'dij_ossz':  None,
        'mindossz':  None,
    }

    # Keresés a fejléc sorban az extra szerepekre
    extra_kws = {
        'anyag_egy':  ['anyag egy', 'anyag egys', 'a:', 'a egys', '∑a', 'anyagár'],
        'dij_egy':    ['díj egy', 'dij egy', 'd:', 'd egys', 'munkadíj egy', 'egység munkadíj', 'egyseg munkadij', 'munkadíj'],
        'anyag_ossz': ['anyag össz', 'anyag össz', 'a ö', '∑a', 'anyag össz'],
        'dij_ossz':   ['díj össz', 'dij össz', 'd ö', '∑d', 'munkadíj össz'],
        'mindossz':   ['mindössz', 'mindösszesen', 'a+d', 'a ö+d ö', '∑a+d', 'osszesen netto', 'a ö+d ö:'],
    }

    for c in range(1, ws.max_column+1):
        v = str(ws.cell(header_row, c).value or '').lower().strip()
        if not v:
            continue
        for role, kws in extra_kws.items():
            if result[role] is None and any(kw in v for kw in kws):
                result[role] = c

    return result

def col_to_cell(col_idx, row):
    """Oszlop index + sor → cellacím pl. (3, 1) → 'C1'"""
    if col_idx is None:
        return ''
    return f'{get_column_letter(col_idx)}{row}'

def cell_to_col(cell_ref):
    """Cellacím → oszlop index pl. 'C1' → 3"""
    if not cell_ref or str(cell_ref).strip() == '':
        return None
    from openpyxl.utils import column_index_from_string
    import re
    m = re.match(r'([A-Za-z]+)(\d+)', str(cell_ref).strip())
    if m:
        return column_index_from_string(m.group(1))
    return None

def generate_mapping_excel(input_path, mapping_path, log_fn=None):
    """
    Beolvassa a bejövő fájlt, minden laphoz meghatározza a javasolt
    cellacím-párosítást, és kiírja egy _MAPPING.xlsx fájlba.
    """
    ext = os.path.splitext(input_path)[1].lower()
    if ext == '.xls':
        wb_src = xls_to_workbook(input_path, log_fn=log_fn)
    else:
        wb_src = load_workbook(input_path)

    SKIP_SHEETS = {'belső', 'főösszesítő', 'foosszesito', 'összesítő', '00_előlap'}
    sheets = [s for s in wb_src.sheetnames if s.lower() not in SKIP_SHEETS]

    wb_map = Workbook()
    wb_map.remove(wb_map.active)

    # Stílusok
    HDR_FILL   = PatternFill("solid", fgColor="1F497D")
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
    GRP_FILL_J = PatternFill("solid", fgColor="DCE6F1")  # Javasolt csoport
    GRP_FILL_E = PatternFill("solid", fgColor="E2EFDA")  # Elfogadott csoport
    ROLE_FONT  = Font(bold=True, size=9)
    INPUT_FILL = PatternFill("solid", fgColor="FFFF99")
    THIN = Side(style='thin')
    BRD  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws_map = wb_map.create_sheet("Párosítás")

    # Fejléc sorok
    ws_map.row_dimensions[1].height = 22
    ws_map.row_dimensions[2].height = 22

    # A1: Munkalap neve
    ws_map.merge_cells('A1:A2')
    c = ws_map['A1']; c.value = 'Munkalap neve'; c.font = HDR_FONT
    c.fill = HDR_FILL; c.alignment = CENTER; c.border = BRD
    ws_map.column_dimensions['A'].width = 28

    # Csoportfejlécek: Javasolt cella (B-G), Elfogadott cella (H-M)
    roles_count = len(BELSO_ROLES)

    # Javasolt csoport
    j_start, j_end = 2, 2 + roles_count - 1
    ws_map.merge_cells(start_row=1, start_column=j_start,
                        end_row=1,   end_column=j_end)
    c = ws_map.cell(1, j_start, value='Javasolt cella')
    c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = BRD

    # Elfogadott csoport
    e_start = j_end + 1
    e_end   = e_start + roles_count - 1
    ws_map.merge_cells(start_row=1, start_column=e_start,
                        end_row=1,   end_column=e_end)
    c = ws_map.cell(1, e_start, value='Elfogadott cella')
    c.font = HDR_FONT
    c.fill = PatternFill("solid", fgColor="1E6B3C")
    c.alignment = CENTER; c.border = BRD

    # Szerepek fejlécei (2. sor)
    for i, (role, label) in enumerate(BELSO_ROLES):
        # Javasolt
        c = ws_map.cell(2, j_start + i, value=label)
        c.font = ROLE_FONT; c.fill = GRP_FILL_J; c.alignment = CENTER; c.border = BRD
        ws_map.column_dimensions[get_column_letter(j_start+i)].width = 14
        # Elfogadott
        c = ws_map.cell(2, e_start + i, value=label)
        c.font = ROLE_FONT; c.fill = GRP_FILL_E; c.alignment = CENTER; c.border = BRD
        ws_map.column_dimensions[get_column_letter(e_start+i)].width = 14

    # Adatsorok – minden lap egy sor
    data_row = 3
    for sname in sheets:
        ws_src = wb_src[sname]
        mapping = detect_full_mapping(ws_src)
        hr = mapping.get('header_row', 1)

        if log_fn: log_fn(f"  {sname}: fejléc sor={hr}, mapping={mapping}")

        # Munkalap neve
        c = ws_map.cell(data_row, 1, value=sname)
        c.font = Font(bold=True); c.border = BRD; c.alignment = Alignment(vertical='center')

        for i, (role, label) in enumerate(BELSO_ROLES):
            col_idx = mapping.get(role)
            cell_ref = col_to_cell(col_idx, hr) if col_idx else ''

            # Javasolt cella (szürke, olvasható)
            cj = ws_map.cell(data_row, j_start + i, value=cell_ref)
            cj.fill = GRP_FILL_J; cj.alignment = CENTER; cj.border = BRD

            # Elfogadott cella (sárga, szerkeszthető – előre feltöltve a javaslattal)
            ce = ws_map.cell(data_row, e_start + i, value=cell_ref)
            ce.fill = INPUT_FILL; ce.alignment = CENTER; ce.border = BRD

        ws_map.row_dimensions[data_row].height = 20
        data_row += 1

    # Útmutató sor
    ws_map.cell(data_row + 1, 1,
        value="ℹ Módosítsd az 'Elfogadott cella' oszlopokat (pl. C4, D4), majd futtasd újra a konverziót.").font = Font(italic=True, color="595959")

    # Rögzítés
    ws_map.freeze_panes = 'B3'

    wb_map.save(mapping_path)
    if log_fn: log_fn(f"  → Mapping fájl mentve: {os.path.basename(mapping_path)}")
    return sheets

def read_mapping_excel(mapping_path):
    """
    Beolvassa a _MAPPING.xlsx fájlt és visszaadja a felhasználó által
    jóváhagyott cellacím-párosításokat lapnév szerint.
    """
    wb = load_workbook(mapping_path, data_only=True)
    ws = wb.active if 'Párosítás' not in wb.sheetnames else wb['Párosítás']

    # Elfogadott csoport oszlopainak megkeresése (2. sor fejléc)
    role_cols = {}  # role → col_index
    roles_order = [r for r, _ in BELSO_ROLES]
    in_elfogadott = False

    for c in range(1, ws.max_column+1):
        h1 = str(ws.cell(1, c).value or '').strip()
        h2 = str(ws.cell(2, c).value or '').strip()
        if 'Elfogadott' in h1:
            in_elfogadott = True
        if in_elfogadott and h2:
            # Megkeresi a role-t a label alapján
            for role, label in BELSO_ROLES:
                if label.lower() in h2.lower() and role not in role_cols:
                    role_cols[role] = c

    # Adatsorok: 3. sortól
    result = {}
    for r in range(3, ws.max_row+1):
        sname = ws.cell(r, 1).value
        if not sname or str(sname).startswith('ℹ'):
            continue
        sheet_mapping = {}
        for role, col in role_cols.items():
            cell_ref = ws.cell(r, col).value
            if cell_ref:
                sheet_mapping[role] = cell_to_col(str(cell_ref))
            else:
                sheet_mapping[role] = None
        result[str(sname)] = sheet_mapping

    return result



def detect_header_row(ws):
    """
    Megkeresi a fejléc sort — ahol TÖBB oszlopban is fejléc-kulcsszó van.
    Felismeri a rövidített fejléceket is: m/e/a:/d: (Szépvölgyi típus).
    """
    # Normál szöveges fejlécek
    header_kws = [
        'mennyiség','menny.','tétel','megnevez','azonosít',
        'egység','egys.','qty','ssz','no.'
    ]
    # Rövidített numerikus fejléc cellák (Szépvölgyi: sor ahol C=tétel szövege, D=m, E=e)
    short_kws = ['tétel szövege', 'tétel', 'megnevez']

    best_row, best_count = 1, 0
    for r in range(1, min(15, ws.max_row+1)):
        count = 0
        has_short_leiras = False
        for c in range(1, ws.max_column+1):
            v = str(ws.cell(r, c).value or '').lower().strip()
            if any(kw in v for kw in header_kws):
                count += 1
            if any(kw in v for kw in short_kws):
                has_short_leiras = True
            # Rövidített fejléc: 'm' (mennyiség) és 'e' (egység) egymás mellett
            if v == 'm' and c < ws.max_column:
                next_v = str(ws.cell(r, c+1).value or '').lower().strip()
                if next_v == 'e':
                    count += 2  # Erős jelzés: m+e egymás mellett = fejléc sor
        if has_short_leiras:
            count += 1
        if count > best_count:
            best_count = count
            best_row = r
    return best_row

def detect_column_mapping(ws, header_row):
    """
    Megkeresi az oszlopok pozícióját a fejléc sor alapján.
    Ha nincs valódi fejléc, az adatsorok mintájából következtet.
    Visszaad: {role: col_index}
    """
    mapping = {'ssz': None, 'leiras': None, 'mennyiseg': None, 'egyseg': None}

    keywords = {
        'ssz':       ['ssz', 'sorsz', 'azonosít', 'tételszám'],
        'leiras':    ['megnevez', 'tétel', 'specif', 'leírás'],
        'mennyiseg': ['mennyiség', 'menny.', 'qty', 'darab', 'mennyis'],
        'egyseg':    ['egység', 'egys.', 'unit', 'mértékegység'],
    }
    short_exact = {
        'mennyiseg': ['m'],
        'egyseg':    ['e'],
    }

    for c in range(1, ws.max_column+1):
        v = str(ws.cell(header_row, c).value or '').lower().strip()
        if not v:
            continue
        for role, kws in keywords.items():
            if mapping[role] is None and any(kw in v for kw in kws):
                mapping[role] = c
        for role, exacts in short_exact.items():
            if mapping[role] is None and v in exacts:
                neighbors = [str(ws.cell(header_row, cc).value or '').lower().strip()
                             for cc in range(max(1,c-2), min(ws.max_column+1, c+3))]
                if any(nb in ['e','m','a:','d:','a ö:','d ö:'] for nb in neighbors):
                    mapping[role] = c

    # Ha a fejléc alapján nem sikerült meghatározni a mappingot,
    # az adatsorok mintájából következtetünk (pl. AGORA típus: A=ssz, B=leírás, C=menny, D=egység)
    if mapping['mennyiseg'] is None or mapping['leiras'] is None:
        data_start = header_row + 1
        for r in range(data_start, min(data_start+8, ws.max_row+1)):
            row_vals = {c: ws.cell(r, c).value for c in range(1, ws.max_column+1)}
            non_empty = {c: v for c, v in row_vals.items() if v is not None}
            if not non_empty:
                continue

            # Keresünk egy sort ahol: valahol egész szám (ssz), valahol szöveg (leírás),
            # valahol szám (mennyiség), valahol rövid szöveg (egység)
            int_cols, str_cols, float_cols, short_str_cols = [], [], [], []
            for c, v in non_empty.items():
                sv = str(v).strip()
                try:
                    n = int(sv)
                    int_cols.append((c, n))
                except:
                    try:
                        float(sv.replace(',','.'))
                        float_cols.append(c)
                    except:
                        if len(sv) <= 6 and sv.replace(' ','').isalpha():
                            short_str_cols.append(c)
                        elif len(sv) > 6:
                            str_cols.append(c)

            if int_cols and str_cols:
                # A legkisebb int col = ssz, a leghosszabb str col = leírás
                if mapping['ssz'] is None:
                    mapping['ssz'] = min(int_cols, key=lambda x: x[0])[0]
                if mapping['leiras'] is None and str_cols:
                    # Leírás: a leghosszabb szövegű oszlop
                    mapping['leiras'] = max(str_cols, key=lambda c:
                        len(str(ws.cell(r, c).value or '')))
                # Mennyiség: az ssz után következő numerikus oszlop
                if mapping['mennyiseg'] is None:
                    ssz_col = mapping['ssz'] or 1
                    leiras_col = mapping['leiras'] or 2
                    num_candidates = [c for c in (int_cols + [(c,0) for c in float_cols])
                                     if c[0] != ssz_col and c[0] != leiras_col]
                    if num_candidates:
                        # A leírás utáni első szám = mennyiség
                        after_leiras = [c for c in num_candidates if c[0] > leiras_col]
                        if after_leiras:
                            mapping['mennyiseg'] = min(after_leiras, key=lambda x: x[0])[0]
                        else:
                            mapping['mennyiseg'] = min(num_candidates, key=lambda x: x[0])[0]
                # Egység: a mennyiség utáni rövid szöveg
                if mapping['egyseg'] is None and short_str_cols and mapping['mennyiseg']:
                    after_menny = [c for c in short_str_cols if c > mapping['mennyiseg']]
                    if after_menny:
                        mapping['egyseg'] = min(after_menny)
                break

    # Végső ellenőrzés: ha A oszlop szekvenciális sorszámokat tartalmaz → ssz=1
    a_nums = []
    for r in range(header_row+1, min(header_row+6, ws.max_row+1)):
        a_val = ws.cell(r, 1).value
        if a_val is not None:
            try: a_nums.append(int(str(a_val).strip()))
            except: break
    if len(a_nums) >= 2 and a_nums == list(range(a_nums[0], a_nums[0]+len(a_nums))):
        mapping['ssz'] = 1
        if mapping['leiras'] in (None, 1):
            mapping['leiras'] = 2
        if not mapping['mennyiseg'] or mapping['mennyiseg'] <= 2:
            mapping['mennyiseg'] = 3
        if not mapping['egyseg'] or mapping['egyseg'] <= 3:
            mapping['egyseg'] = 4

    return mapping

def parse_sheet(ws_src, log_fn=None, user_mapping=None):
    """
    Beolvassa a forrás lapot, visszaadja a tételek listáját.
    Felismeri automatikusan:
      - Normál struktúra (Mennyiség, Egység, Tétel megnevezése oszlopok)
      - Rövidített struktúra (m, e, a:, d: — pl. Szépvölgyi típus)
      - Több fejezetes struktúra (sötét fejezet sorok, alfejezet szövegek)
    """
    header_row = detect_header_row(ws_src)
    mapping    = detect_column_mapping(ws_src, header_row)

    if log_fn:
        log_fn(f"    Fejléc sor: {header_row}")
        log_fn(f"    Oszlop mapping: {mapping}")

    # Felhasználói párosítás felülírja az automatikus detektálást
    if user_mapping:
        if user_mapping.get('mennyiseg'): mapping['mennyiseg'] = user_mapping['mennyiseg']
        if user_mapping.get('egyseg'):    mapping['egyseg']    = user_mapping['egyseg']
        if user_mapping.get('anyag_egy'):
            mapping['anyag_egy'] = user_mapping['anyag_egy']
        if log_fn: log_fn(f"    Felhasználói mapping alkalmazva: {user_mapping}")

    col_ssz  = mapping.get('ssz')    or 1
    col_b    = mapping.get('leiras') or 2
    col_menn = mapping.get('mennyiseg') or 3
    col_egy  = mapping.get('egyseg') or 4

    # Ellenőrizzük: a header_row maga is adatsor-e?
    # Ha a header_row-ban ssz és mennyiség is van, az adatok ott kezdődnek (nincs valódi fejléc)
    hr_ssz  = ws_src.cell(header_row, col_ssz).value
    hr_menn = ws_src.cell(header_row, col_menn).value
    try:
        int(str(hr_ssz or '').strip())
        float(str(hr_menn or '').replace(',','.'))
        data_start = header_row  # a fejléc sor maga is adatsor!
    except:
        data_start = header_row + 1

    items = []
    pending = []

    for r in range(data_start, ws_src.max_row + 1):
        mennyiseg   = ws_src.cell(r, col_menn).value
        leiras_cell = ws_src.cell(r, col_b).value
        egyseg      = ws_src.cell(r, col_egy).value
        ssz_cell    = ws_src.cell(r, col_ssz).value

        # Teljesen üres sor kihagyás
        if mennyiseg is None and leiras_cell is None and ssz_cell is None:
            continue

        # Összesítő / opciós / alfejezet sorok kihagyása
        leiras_str = str(leiras_cell or '').strip()
        if any(w in leiras_str.lower() for w in SKIP_WORDS):
            continue

        # Fejezet cím sor (A = nagy egész pl. 17, nincs mennyiség)
        if ssz_cell is not None and mennyiseg is None:
            try:
                n = int(str(ssz_cell).strip())
                if n > 100:  # Fejezet azonosító (pl. 101000, 17, stb.) — kihagyás, NEM pending
                    continue
            except:
                pass

        # Alfejezet szöveg (csak C/B oszlopban szöveg, nincs menny)
        if mennyiseg is None and leiras_str and not str(leiras_str).startswith('='):
            pending.append(leiras_str)
            continue

        # Valódi adatsor: van mennyiség
        if mennyiseg is not None:
            # Sorszám
            try:
                ssz = int(str(ssz_cell).strip()) if ssz_cell is not None else len(items)+1
            except:
                ssz = len(items)+1

            # Mennyiség
            try:
                qty = float(str(mennyiseg).replace(',', '.'))
            except:
                qty = mennyiseg

            # Leírás összerakása: pending + aktuális sor szövege
            parts = pending + ([leiras_str] if leiras_str else [])
            pending = []

            items.append({
                'ssz':       ssz,
                'leiras':    '\n'.join(p for p in parts if p),
                'mennyiseg': qty,
                'egyseg':    str(egyseg).strip() if egyseg else '',
            })

    return items

def copy_sheet(wb_src, sname, wb_out, out_name):
    """Átmásolja az eredeti munkalapot az output workbook-ba (értékek + formátum)."""
    import copy as _copy
    ws_src = wb_src[sname]
    ws_new = wb_out.create_sheet(title=out_name)

    # Cellák másolása
    for row in ws_src.iter_rows():
        for cell in row:
            new_cell = ws_new.cell(row=cell.row, column=cell.column)
            # Érték (képleteket értékként másoljuk)
            if cell.data_type == 'f':
                new_cell.value = cell.value  # képlet marad
            else:
                new_cell.value = cell.value
            # Formátum
            if cell.has_style:
                new_cell.font      = cell.font.copy()
                new_cell.fill      = cell.fill.copy()
                new_cell.border    = cell.border.copy()
                new_cell.alignment = cell.alignment.copy()
                new_cell.number_format = cell.number_format

    # Sormagasságok
    for rn, rd in ws_src.row_dimensions.items():
        ws_new.row_dimensions[rn].height = rd.height

    # Oszlopszélességek
    for cn, cd in ws_src.column_dimensions.items():
        ws_new.column_dimensions[cn].width = cd.width

    # Összevont cellák
    for mc in ws_src.merged_cells.ranges:
        ws_new.merge_cells(str(mc))

    return ws_new


def append_belso_columns(ws, mapping, data_rows):
    """
    Hozzáadja a BELSŐ oszlopokat az eredeti lap jobb oldalára.
    mapping: {role: col_idx} — forrás oszlopok
    data_rows: [(row_number, ssz)] — adatsorok sorszámai
    """
    # Az első üres oszlop után kezdünk, egy elválasztó oszlopot hagyva
    insert_col = ws.max_column + 2

    # Fejléc sor a mapping-ból
    hr = mapping.get('header_row', 1)

    # Elválasztó oszlop fejléce
    sep = ws.cell(hr, insert_col - 1)
    sep.value = '|'
    sep.font = Font(bold=True, color='AAAAAA')

    # BELSŐ fejléc cím
    title_cell = ws.cell(hr - 1 if hr > 1 else hr, insert_col)
    title_cell.value = 'BELSŐ KALKULÁCIÓ'
    title_cell.font = Font(bold=True, color='FFFFFF')
    title_cell.fill = PatternFill('solid', fgColor='1A3A5C')
    if hr > 1:
        try:
            ws.merge_cells(start_row=hr-1, start_column=insert_col,
                           end_row=hr-1,   end_column=insert_col+12)
        except Exception:
            pass

    # BELSŐ fejlécek kiírása a header sorba
    belso_cols = [
        ('E', 'Anyag be',          '#FFFF00'),
        ('F', 'EURO',              '#FFFF00'),
        ('H', 'Kedv.',             '#FFFF00'),
        ('I', 'Kedv. Ár',          None),
        ('J', 'Ax kód',            '#FFFF00'),
        ('K', 'Ax',                None),
        ('L', 'Anyag egysar',      None),
        ('M', 'No.ido',            '#FFFF00'),
        ('N', 'Egyseg munkadij',   None),
        ('O', 'Anyag osszesen',    None),
        ('P', 'Munkadij osszesen', None),
        ('Q', 'Osszesen',          None),
    ]

    for i, (role_letter, label, fill_hex) in enumerate(belso_cols):
        c = insert_col + i
        cell = ws.cell(hr, c, value=label)
        cell.font = Font(bold=True, size=8)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = BORDER
        if fill_hex:
            cell.fill = PatternFill('solid', fgColor=fill_hex.lstrip('#'))
        ws.column_dimensions[get_column_letter(c)].width = 11

    # Adatsorok: BELSŐ képletek hozzáadása
    belso_offset = insert_col  # E oszlop pozíciója az output lapon

    def bc(letter, r):
        """BELSŐ oszlop abszolút cellacím az output lapon"""
        role_map = {
            'E': 0, 'F': 1, 'H': 2, 'I': 3, 'J': 4,
            'K': 5, 'L': 6, 'M': 7, 'N': 8, 'O': 9, 'P': 10, 'Q': 11
        }
        return get_column_letter(belso_offset + role_map[letter]) + str(r)

    # Mennyiség forrásoszlop betűje
    menny_col = mapping.get('mennyiseg')
    menny_letter = get_column_letter(menny_col) if menny_col else 'C'

    for row_num in data_rows:
        r = row_num
        E = belso_offset + 0
        F = belso_offset + 1
        H = belso_offset + 2
        I = belso_offset + 3
        J = belso_offset + 4
        K = belso_offset + 5
        L = belso_offset + 6
        M = belso_offset + 7
        N = belso_offset + 8
        O = belso_offset + 9
        P = belso_offset + 10
        Q = belso_offset + 11

        El = get_column_letter(E)
        Fl = get_column_letter(F)
        Hl = get_column_letter(H)
        Il = get_column_letter(I)
        Jl = get_column_letter(J)
        Kl = get_column_letter(K)
        Ll = get_column_letter(L)
        Ml = get_column_letter(M)
        Nl = get_column_letter(N)
        Ol = get_column_letter(O)
        Pl = get_column_letter(P)
        Ql = get_column_letter(Q)

        def c(col, val, fmt=None, border=True):
            cell = ws.cell(r, col, value=val)
            if fmt: cell.number_format = fmt
            if border: cell.border = BORDER
            return cell

        c(E, 0, '#,##0 "Ft"')                                          # Anyag be
        c(F, None, '#,##0.0 "€"')                                       # EURO
        c(H, 0, '0%')                                                   # Kedv.
        c(I, f'=IF({Fl}{r}>0,{Fl}{r}*BELSŐ!$F$3,{El}{r})*(1-{Hl}{r})',
          '#,##0 "Ft"')                                                 # Kedv. Ár
        c(J, 'Anyag')                                                   # Ax kód
        c(K, f'=VLOOKUP({Jl}{r},BELSŐ!C:D,2,FALSE)', '0%')             # Ax
        c(L, f'={Il}{r}*(1+{Kl}{r})', '#,##0 "Ft"')                    # Anyag egységár
        c(M, 0, '#,##0.##')                                             # No.idő
        c(N, f'={Ml}{r}*BELSŐ!$D$3', '#,##0 "Ft"')                     # Egység munkadíj
        c(O, f'={menny_letter}{r}*{Ll}{r}', '#,##0 "Ft"')              # Anyag összesen
        c(P, f'={menny_letter}{r}*{Nl}{r}', '#,##0 "Ft"')              # Munkadíj összesen
        c(Q, f'={Ol}{r}+{Pl}{r}', '#,##0 "Ft"')                        # Összesen

        # Feltételes formázás - egyszerű háttér az input cellákra
        ws.cell(r, E).fill = make_cf_fill('FFFFBFBF')  # Anyag be
        ws.cell(r, M).fill = make_cf_fill('FFFFBFBF')  # No.idő

        # Eredeti oszlopok atahhivatkozasa a BELSO kalkulalt ertekekre
        anyag_egy_col = mapping.get('anyag_egy')
        dij_egy_col   = mapping.get('dij_egy')
        if anyag_egy_col:
            orig = ws.cell(r, anyag_egy_col)
            orig.value = f'={Ll}{r}'      # BELSO anyag egységár kalkulált
            orig.number_format = '#,##0 "Ft"'
        # Ha a dij_egy mapping nem talalta meg, pozicio alapjan keresunk:
        # az anyag_egy utan 2 oszloppal szokott lenni az egyseg munkadij
        if dij_egy_col is None and anyag_egy_col:
            hr = mapping.get('header_row', 1)
            for offset in [2, 1, 3]:
                candidate = anyag_egy_col + offset
                hdr_val = str(ws.cell(hr, candidate).value or '').lower()
                if any(kw in hdr_val for kw in ['munkadij', 'munkadíj', 'dij', 'díj']):
                    dij_egy_col = candidate
                    break
        if dij_egy_col:
            orig = ws.cell(r, dij_egy_col)
            orig.value = f'={Nl}{r}'      # BELSO egység munkadíj
            orig.number_format = '#,##0 "Ft"'

    # Összesítő sor az adatsorok után
    if data_rows:
        last_data = max(data_rows)
        sum_r = last_data + 2

        # Keresés: van-e már "Összesen" sor az eredetiben?
        for r in range(last_data+1, min(last_data+5, ws.max_row+1)):
            v = str(ws.cell(r, 1).value or '') + str(ws.cell(r, 2).value or '') + str(ws.cell(r, 3).value or '')
            if 'összesen' in v.lower() or 'osszesen' in v.lower():
                sum_r = r
                break

        first_data = min(data_rows)
        for col_offset, label in [(9,'Anyag össz.'),(10,'Munkadíj össz.'),(11,'Összesen')]:
            col = belso_offset + col_offset
            cl  = get_column_letter(col)
            cell = ws.cell(sum_r, col,
                           value=f'=SUM({cl}{first_data}:{cl}{last_data})')
            cell.font = Font(bold=True)
            cell.border = Border(bottom=Side(style='double'))
            cell.number_format = '#,##0 "Ft"'

    # Feltételes formázás az Anyag be és No.idő oszlopokra
    E_col = get_column_letter(belso_offset)
    M_col = get_column_letter(belso_offset + 7)
    first_r = min(data_rows) if data_rows else 2
    last_r  = max(data_rows) if data_rows else 2
    add_conditional_formatting_cols(ws, E_col, M_col, first_r, last_r)

    return insert_col


def add_conditional_formatting_cols(ws, e_col, m_col, first_r, last_r):
    """Feltételes formázás két adott oszlopra."""
    EQ  = DifferentialStyle(font=Font(color='FF9C0006'), fill=make_cf_fill('FFFFBFBF'))
    LT  = DifferentialStyle(font=Font(color='FF9C0006'), fill=make_cf_fill('FF008000'))
    GT  = DifferentialStyle(font=Font(color='FF000000'), fill=make_cf_fill('FFBFFFBF'))
    GRN = DifferentialStyle(font=Font(color='FF000000'), fill=make_cf_fill('FFBFFFBF'))
    for col in (e_col, m_col):
        rng = f'{col}{first_r}:{col}{last_r}'
        ws.conditional_formatting.add(rng, Rule(type='cellIs', operator='equal',       formula=['0'], dxf=EQ,  priority=4))
        ws.conditional_formatting.add(rng, Rule(type='cellIs', operator='lessThan',    formula=['0'], dxf=LT,  priority=5))
        ws.conditional_formatting.add(rng, Rule(type='cellIs', operator='greaterThan', formula=['0'], dxf=GT,  priority=6))
    f_rng = f'{e_col}{first_r}:{e_col}{last_r}'
    e_next = get_column_letter(column_index_from_string(e_col) + 1)
    ws.conditional_formatting.add(f_rng,
        Rule(type='expression', formula=[f'{e_next}{first_r}>0'], dxf=GRN, priority=1))


def find_data_rows(ws_src, mapping):
    """Megkeresi az adatsorok sorszámait (ahol van mennyiség)."""
    col_menn = mapping.get('mennyiseg') or 3
    header_row = mapping.get('header_row', 1)
    data_rows = []
    for r in range(header_row + 1, ws_src.max_row + 1):
        val = ws_src.cell(r, col_menn).value
        if val is None:
            continue
        try:
            float(str(val).replace(',', '.'))
            data_rows.append(r)
        except (ValueError, TypeError):
            pass
    return data_rows

def write_sheet(wb_out, sheet_name, items):
    ws = wb_out.create_sheet(title=sheet_name)
    set_header(ws)
    for i, item in enumerate(items, start=1):
        write_data_row(ws, i+1, i, item['leiras'], item['mennyiseg'], item['egyseg'])
        lines = item['leiras'].count('\n') + 1
        ws.row_dimensions[i+1].height = max(18, min(lines * 15, 100))
    data_last = len(items) + 1
    # Összesítő
    sum_row = data_last + 2
    ws.cell(data_last + 1, 2, value='Összesen:').font = BOLD
    ws.cell(sum_row, 13, value='Összesen nettó:').font = BOLD
    ws.cell(sum_row, 13).alignment = Alignment(horizontal='right')
    for col_idx in [14, 15, 16, 17]:
        cl = get_column_letter(col_idx)
        cell = ws.cell(sum_row, col_idx, value=f'=SUM({cl}2:{cl}{data_last})')
        cell.font = BOLD
        cell.border = Border(bottom=Side(style='double'))
        cell.number_format = '#,##0 "Ft"'
    add_conditional_formatting(ws, data_last)
    # Oszlopszélességek
    for ci, w in {1:6,2:52,3:10,4:9,5:12,6:10,7:8,8:8,9:12,10:9,11:7,12:14,13:8,14:14,15:14,16:16,17:12}.items():
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'B2'
    return data_last

def create_belso_sheet(wb):
    ws = wb.create_sheet('BELSŐ')
    ws['B1'] = 'Szerelési óradíjak:'; ws['B1'].font = BOLD
    ws['D2'] = 'FO'; ws['F2'] = 'EURO Árfolyam:'
    for i, (nev, dij) in enumerate([('Szerelés',8150),('Légtechnika',7900),
        ('Légtechnika kör idom',7500),('Légtechnika négyszög',7500),
        ('Légtechnika Spiko cső',7500),('BONTÁS',7370)], start=3):
        ws.cell(i,2,nev); ws.cell(i,3,'Ft/óra'); ws.cell(i,4,dij)
    ws['F3'] = 375; ws['G3'] = 'EURO'
    for i, (k,v) in enumerate([('Anyag',0.18),('Bontás',0.18),('Szigetelés',0.18),
        ('Rézcső',0.18),('Acélcső',0.18)], start=11):
        ws.cell(i,3,k); ws.cell(i,4,v)
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10

def create_foosszesito(wb, sheet_names, project_title=""):
    ws = wb.create_sheet('Főösszesítő', 0)
    LTBLUE = PatternFill("solid", fgColor="BDD7EE")
    ws.merge_cells('B2:F2')
    ws['B2'] = project_title or 'GÉPÉSZ FŐÖSSZESÍTŐ'
    ws['B2'].font = Font(bold=True, size=13)
    ws['B2'].alignment = CENTER
    headers = ['Fejezet','Megnevezés','Anyag összesen','Munkadíj összesen','Összesen nettó']
    for col, h in enumerate(headers, start=2):
        cell = ws.cell(8, col, value=h)
        cell.font = BOLD; cell.fill = LTBLUE; cell.border = BORDER; cell.alignment = CENTER
    row = 9
    for sname in sheet_names:
        ws.cell(row,2,value=sname).border = BORDER
        ws.cell(row,3,value=sname).border = BORDER
        ws.cell(row,4,value=f"=IFERROR(SUM('{sname}'!O2:O500),0)").border = BORDER
        ws.cell(row,5,value=f"=IFERROR(SUM('{sname}'!P2:P500),0)").border = BORDER
        ws.cell(row,6,value=f"=IFERROR(SUM('{sname}'!Q2:Q500),0)").border = BORDER
        for c in range(4,7): ws.cell(row,c).number_format = '#,##0 "Ft"'
        row += 1
    for c_idx, label, formula in [
        (3,'ÖSSZESEN (NETTÓ)', None),(4,None,f'=SUM(D9:D{row-1})'),
        (5,None,f'=SUM(E9:E{row-1})'),(6,None,f'=SUM(F9:F{row-1})')]:
        if label: ws.cell(row+1,c_idx,value=label).font = BOLD
        if formula:
            c = ws.cell(row+1,c_idx,value=formula); c.font = BOLD; c.number_format = '#,##0 "Ft"'
    ws.cell(row+2,3,'27% ÁFA')
    ws.cell(row+2,6,f'=F{row+1}*0.27').number_format = '#,##0 "Ft"'
    ws.cell(row+3,3,'MINDÖSSZESEN (BRUTTÓ)').font = BOLD
    ws.cell(row+3,6,f'=F{row+1}+F{row+2}').font = BOLD
    ws.cell(row+3,6).number_format = '#,##0 "Ft"'
    for col,w in [(2,8),(3,28),(4,16),(5,18),(6,16)]:
        ws.column_dimensions[get_column_letter(col)].width = w

def xls_to_workbook(xls_path, log_fn=None):
    """Régi .xls fájl beolvasása xlrd-vel, visszaad egy openpyxl Workbook-ot."""
    try:
        import xlrd
    except ImportError:
        raise RuntimeError(
            "A régi .xls formátum olvasásához az xlrd csomag szükséges.\n"
            "Telepítés: pip install xlrd\n"
            "Vagy mentsd el a fájlt .xlsx formátumban Excelben (Mentés másként)."
        )
    if log_fn: log_fn("  -> Régi .xls formátum - konvertálás...")
    xls_wb = xlrd.open_workbook(xls_path, formatting_info=False)
    new_wb = Workbook()
    new_wb.remove(new_wb.active)
    for sheet_name in xls_wb.sheet_names():
        xls_ws = xls_wb.sheet_by_name(sheet_name)
        new_ws = new_wb.create_sheet(title=sheet_name)
        for ri in range(xls_ws.nrows):
            for ci in range(xls_ws.ncols):
                cell = xls_ws.cell(ri, ci)
                if cell.ctype == 1:    val = cell.value
                elif cell.ctype == 2:
                    val = int(cell.value) if cell.value == int(cell.value) else cell.value
                elif cell.ctype == 4:  val = bool(cell.value)
                elif cell.ctype in (0, 5): val = None
                else:                  val = cell.value
                if val is not None:
                    new_ws.cell(row=ri+1, column=ci+1, value=val)
    if log_fn: log_fn(f"  -> .xls beolvasva: {len(xls_wb.sheet_names())} lap")
    return new_wb

def convert_file(input_path, output_path, progress_cb=None, log_cb=None, mapping_path=None, user_mappings=None):
    """
    Fő konverziós függvény.
    progress_cb(float 0..1), log_cb(str)
    """
    def log(msg):
        if log_cb: log_cb(msg)

    log(f"Betöltés: {os.path.basename(input_path)}")
    ext = os.path.splitext(input_path)[1].lower()
    if ext == '.xls':
        wb_src = xls_to_workbook(input_path, log_fn=log)
    else:
        wb_src = load_workbook(input_path)
    sheets = [s for s in wb_src.sheetnames
              if s.lower() not in ('belső', 'főösszesítő', 'foosszesito', 'összesítő', '00_előlap')]
    log(f"Talált lapok: {sheets}")

    # Párosítás
    if user_mappings is None:
        user_mappings = {}
        if mapping_path and os.path.exists(mapping_path):
            try:
                user_mappings = read_mapping_excel(mapping_path)
                log(f"  Párosítás betöltve fájlból: {len(user_mappings)} lap")
            except Exception as e:
                log(f"  ! Párosítás hiba: {e}")

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    out_sheet_names = []
    for i, sname in enumerate(sheets):
        log(f"\n→ Feldolgozás: {sname}")
        ws_src = wb_src[sname]

        # Mapping detektálás (auto + felhasználói felülírás)
        full_map = detect_full_mapping(ws_src)
        user_map = user_mappings.get(sname, {})
        if user_map:
            full_map.update({k: v for k, v in user_map.items() if v is not None})

        # Adatsorok megkeresése
        data_rows = find_data_rows(ws_src, full_map)
        log(f"  {len(data_rows)} adatsor")
        if not data_rows:
            log(f"  ! Üres lap, kihagyva")
            continue

        out_name = sname[:31]

        # Eredeti lap másolása + BELSŐ oszlopok hozzáadása
        ws_out = copy_sheet(wb_src, sname, wb_out, out_name)
        append_belso_columns(ws_out, full_map, data_rows)

        # BELSŐ referencia lap neve frissítése ha szükséges
        out_sheet_names.append(out_name)
        if progress_cb:
            progress_cb((i+1) / (len(sheets)+1))

    if out_sheet_names:
        create_foosszesito(wb_out, out_sheet_names,
                           project_title=os.path.splitext(os.path.basename(input_path))[0])
        create_belso_sheet(wb_out)

    wb_out.save(output_path)
    log(f"\n✓ Kész! → {os.path.basename(output_path)}")
    if progress_cb: progress_cb(1.0)
    return len(out_sheet_names)


# ── GUI ───────────────────────────────────────────────────────────────────────

class App(TkinterDnD.Tk if HAS_DND else tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("KK-Industry – BELSŐ Konverter")
        self.geometry("600x520")
        self.resizable(False, False)
        self.configure(bg="#F5F5F5")
        self._input_path = tk.StringVar()
        self._output_path = tk.StringVar()
        self._build_ui()

    def _build_ui(self):
        # Fejléc
        hdr = tk.Frame(self, bg="#1A3A5C", height=56)
        hdr.pack(fill='x')
        hdr.pack_propagate(False)
        tk.Label(hdr, text="⚙  BELSŐ Költségvetés Konverter",
                 bg="#1A3A5C", fg="white",
                 font=("Segoe UI", 13, "bold")).pack(side='left', padx=20, pady=14)
        tk.Label(hdr, text="KK-Industry Kft.",
                 bg="#1A3A5C", fg="#7FB3D3",
                 font=("Segoe UI", 9)).pack(side='right', padx=20)

        body = tk.Frame(self, bg="#F5F5F5")
        body.pack(fill='both', expand=True, padx=24, pady=18)

        # Bejövő fájl – Drag & Drop zóna
        self._section(body, "1  Bejövő Excel fájl kiválasztása")

        # Drop zóna keret
        self._drop_frame = tk.Frame(body, bg="#E8F0F7", relief='flat',
                                    highlightthickness=2,
                                    highlightbackground="#AABCCC")
        self._drop_frame.pack(fill='x', pady=(4,4))

        self._drop_label = tk.Label(
            self._drop_frame,
            text="📂  Húzd ide az Excel fájlt (.xlsx / .xls)",
            bg="#E8F0F7", fg="#1A3A5C",
            font=("Segoe UI", 10), pady=14, cursor="hand2"
        )
        self._drop_label.pack(fill='x')
        self._drop_label.bind("<Button-1>", lambda e: self._browse_input())

        # Kötjük a drop eventet ha elérhető
        if HAS_DND:
            self._drop_frame.drop_target_register(DND_FILES)
            self._drop_frame.dnd_bind("<<Drop>>", self._on_drop)
            self._drop_label.drop_target_register(DND_FILES)
            self._drop_label.dnd_bind("<<Drop>>", self._on_drop)

        # Fájlút sor
        row1 = tk.Frame(body, bg="#F5F5F5")
        row1.pack(fill='x', pady=(4,12))
        self._entry_in = tk.Entry(row1, textvariable=self._input_path,
                                  font=("Segoe UI",10), bg="white",
                                  relief='flat', bd=1, highlightthickness=1,
                                  highlightbackground="#CCCCCC")
        self._entry_in.pack(side='left', fill='x', expand=True, ipady=5, padx=(0,8))
        tk.Button(row1, text="Tallózás…", command=self._browse_input,
                  font=("Segoe UI",9), bg="#E8EDF2", relief='flat',
                  padx=10, pady=5, cursor='hand2').pack(side='right')

        # Kimeneti fájl
        self._section(body, "2  Mentési hely")
        row2 = tk.Frame(body, bg="#F5F5F5")
        row2.pack(fill='x', pady=(4,12))
        self._entry_out = tk.Entry(row2, textvariable=self._output_path,
                                   font=("Segoe UI",10), bg="white",
                                   relief='flat', bd=1, highlightthickness=1,
                                   highlightbackground="#CCCCCC")
        self._entry_out.pack(side='left', fill='x', expand=True, ipady=5, padx=(0,8))
        tk.Button(row2, text="Mentés…", command=self._browse_output,
                  font=("Segoe UI",9), bg="#E8EDF2", relief='flat',
                  padx=10, pady=5, cursor='hand2').pack(side='right')

        # Progress
        self._section(body, "3  Konverzió")
        self._progress = ttk.Progressbar(body, length=552, mode='determinate')
        self._progress.pack(fill='x', pady=(4,4))

        # Log
        log_frame = tk.Frame(body, bg="#F5F5F5")
        log_frame.pack(fill='both', expand=True, pady=(4,12))
        self._log = tk.Text(log_frame, height=8, font=("Consolas",9),
                            bg="#1E1E1E", fg="#D4D4D4", relief='flat',
                            wrap='word', state='disabled')
        scroll = tk.Scrollbar(log_frame, command=self._log.yview)
        self._log.configure(yscrollcommand=scroll.set)
        scroll.pack(side='right', fill='y')
        self._log.pack(side='left', fill='both', expand=True)

        # Gombok – kétlépéses workflow
        btn_frame = tk.Frame(body, bg="#F5F5F5")
        btn_frame.pack(fill='x', pady=(4,0))
        btn_frame.columnconfigure(0, weight=1)
        btn_frame.columnconfigure(1, weight=1)

        self._btn_map = tk.Button(btn_frame,
                              text="1️⃣  PÁROSÍTÁS GENERÁLÁSA",
                              command=self._run_mapping,
                              font=("Segoe UI", 10, "bold"),
                              bg="#1A3A5C", fg="white",
                              activebackground="#122B45", activeforeground="white",
                              relief='flat', padx=12, pady=10, cursor='hand2')
        self._btn_map.grid(row=0, column=0, sticky='ew', padx=(0,4))

        self._btn_conv = tk.Button(btn_frame,
                              text="2️⃣  BELSŐ VERZIÓ GENERÁLÁSA",
                              command=self._run_convert,
                              font=("Segoe UI", 10, "bold"),
                              bg="#1A6B3C", fg="white",
                              activebackground="#145530", activeforeground="white",
                              relief='flat', padx=12, pady=10, cursor='hand2')
        self._btn_conv.grid(row=0, column=1, sticky='ew', padx=(4,0))

        tk.Label(body,
                 text="1️⃣ Párosítás ellenőrzése a programban → 2️⃣ Konvertálás",
                 bg="#F5F5F5", fg="#595959", font=("Segoe UI", 8)).pack(anchor='w', pady=(4,0))

    def _section(self, parent, text):
        tk.Label(parent, text=text, bg="#F5F5F5", fg="#1A3A5C",
                 font=("Segoe UI", 9, "bold")).pack(anchor='w', pady=(4,0))

    def _on_drop(self, event):
        """Drag & drop esemény kezelése."""
        raw = event.data.strip()
        # Windows: {} közé zárva ha szóköz van az útban
        if raw.startswith('{') and raw.endswith('}'):
            raw = raw[1:-1]
        path = raw.split('}')[0].split('{')[-1].strip()
        if not path:
            path = raw
        ext = os.path.splitext(path)[1].lower()
        if ext not in ('.xlsx', '.xls'):
            messagebox.showwarning("Nem támogatott", f"Csak .xlsx vagy .xls fájl fogadható el!\n\n{path}")
            return
        self._set_input(path)
        # Vizuális visszajelzés
        self._drop_frame.configure(highlightbackground="#1A6B3C", bg="#D6EAD9")
        self._drop_label.configure(bg="#D6EAD9", fg="#1A6B3C",
                                    text=f"✓  {os.path.basename(path)}")
        self.after(2000, self._reset_drop_zone)

    def _reset_drop_zone(self):
        self._drop_frame.configure(highlightbackground="#AABCCC", bg="#E8F0F7")
        self._drop_label.configure(bg="#E8F0F7", fg="#1A3A5C",
                                    text="📂  Húzd ide az Excel fájlt (.xlsx / .xls)")

    def _set_input(self, path):
        self._input_path.set(path)
        p = Path(path)
        auto_out = p.parent / f"{p.stem}_BELSŐ.xlsx"
        self._output_path.set(str(auto_out))

    def _browse_input(self):
        path = filedialog.askopenfilename(
            title="Válaszd ki a bejövő Excel fájlt",
            filetypes=[("Excel fájlok", "*.xlsx *.xls"), ("Minden fájl", "*.*")])
        if path:
            self._set_input(path)
            self._drop_label.configure(text=f"✓  {os.path.basename(path)}",
                                        fg="#1A6B3C", bg="#D6EAD9")
            self._drop_frame.configure(bg="#D6EAD9", highlightbackground="#1A6B3C")
            self.after(2000, self._reset_drop_zone)

    def _browse_output(self):
        path = filedialog.asksaveasfilename(
            title="Mentési hely",
            defaultextension=".xlsx",
            filetypes=[("Excel fájlok", "*.xlsx")])
        if path:
            self._output_path.set(path)

    def _log_append(self, msg):
        self._log.configure(state='normal')
        self._log.insert('end', msg + '\n')
        self._log.see('end')
        self._log.configure(state='disabled')

    def _set_progress(self, val):
        self._progress['value'] = val * 100
        self.update_idletasks()

    def _run_mapping(self):
        """1. lépés: párosítási ablak megnyitása a programon belül."""
        inp = self._input_path.get().strip()
        if not inp:
            messagebox.showwarning("Hiányzó fájl", "Válassz bejövő Excel fájlt!")
            return
        if not os.path.exists(inp):
            messagebox.showerror("Fájl nem található", f"Nem létezik:\n{inp}")
            return

        self._btn_map.configure(state='disabled', text="Betöltés…")
        self._clear_log()
        self._log_append("=" * 48)
        self._log_append(f"Fájl elemzése: {os.path.basename(inp)}")

        def load_and_show():
            try:
                ext = os.path.splitext(inp)[1].lower()
                if ext == '.xls':
                    wb_src = xls_to_workbook(inp, log_fn=lambda m: self.after(0, self._log_append, m))
                else:
                    wb_src = load_workbook(inp)

                SKIP = {'belső', 'főösszesítő', 'foosszesito', 'összesítő', '00_előlap'}
                sheets = [s for s in wb_src.sheetnames if s.lower() not in SKIP]

                auto_maps = {}
                for sname in sheets:
                    ws = wb_src[sname]
                    auto_maps[sname] = detect_full_mapping(ws)
                    self.after(0, self._log_append, f"  {sname}: fejléc sor={auto_maps[sname]['header_row']}")

                self.after(0, self._open_mapping_dialog, wb_src, auto_maps)
            except Exception as e:
                self.after(0, self._error, str(e))
                self.after(0, lambda: self._btn_map.configure(state='normal', text="1️⃣  PÁROSÍTÁS"))

        threading.Thread(target=load_and_show, daemon=True).start()

    def _open_mapping_dialog(self, wb_src, auto_maps):
        """Megnyitja a párosítási dialógust a fő szálon."""
        self._btn_map.configure(state='normal', text="1️⃣  PÁROSÍTÁS")
        result = show_mapping_dialog(self, wb_src, auto_maps)
        if result is not None:
            self._user_mappings = result
            self._log_append(f"\n✓ Párosítás elfogadva ({len(result)} lap)")
            for sname, m in result.items():
                parts = [f"{label}={chr(64+m[role]) if m.get(role) else '—'}"
                         for role, label, _ in BELSO_ROLES]
                self._log_append(f"  {sname}: {'  '.join(parts)}")
            self._log_append("\n→ Kattints a 2️⃣ gombra a konverzióhoz!")
        else:
            self._log_append("Párosítás megszakítva.")

    def _run_convert(self):
        """2. lépés: konverzió a párosítás alapján."""
        inp = self._input_path.get().strip()
        out = self._output_path.get().strip()
        if not inp:
            messagebox.showwarning("Hiányzó fájl", "Válassz bejövő Excel fájlt!")
            return
        if not out:
            messagebox.showwarning("Hiányzó útvonal", "Add meg a mentési helyet!")
            return
        user_mappings = getattr(self, '_user_mappings', None)
        if not user_mappings:
            if not messagebox.askyesno("Nincs párosítás",
                "Még nem futtattad a párosítást (1️⃣ gomb).\n\n"
                "Automatikus detektálással folytassuk?"):
                return
        out = self._resolve_writable_path(out)
        self._output_path.set(out)
        self._btn_conv.configure(state='disabled', text="Folyamatban…")
        self._clear_log()
        self._log_append("=" * 48)
        self._log_append(f"Konverzió: {os.path.basename(inp)}")
        self._log_append("=" * 48)
        def worker():
            try:
                count = convert_file(inp, out,
                    user_mappings=user_mappings,
                    progress_cb=lambda v: self.after(0, self._set_progress, v),
                    log_cb=lambda m: self.after(0, self._log_append, m))
                self.after(0, self._done, count, out)
            except Exception as e:
                self.after(0, self._error, str(e))
                self.after(0, lambda: self._btn_conv.configure(state='normal', text="2️⃣  BELSŐ VERZIÓ GENERÁLÁSA"))
        threading.Thread(target=worker, daemon=True).start()

    def _clear_log(self):
        self._log.configure(state='normal')
        self._log.delete('1.0', 'end')
        self._log.configure(state='disabled')

    def _run(self):
        inp = self._input_path.get().strip()
        out = self._output_path.get().strip()
        if not inp:
            messagebox.showwarning("Hiányzó fájl", "Válassz bejövő Excel fájlt!")
            return
        if not out:
            messagebox.showwarning("Hiányzó útvonal", "Add meg a mentési helyet!")
            return
        if not os.path.exists(inp):
            messagebox.showerror("Fájl nem található", f"Nem létezik:\n{inp}")
            return

        # Írhatóság ellenőrzése — ha nem írható, fallback Asztalra / Letöltésekbe
        out = self._resolve_writable_path(out)
        self._output_path.set(out)

        self._btn.configure(state='disabled', text="Folyamatban…")
        self._progress['value'] = 0
        self._log.configure(state='normal')
        self._log.delete('1.0', 'end')
        self._log.configure(state='disabled')
        self._log_append(f"{'='*50}")
        self._log_append(f"Indítás: {datetime.now().strftime('%H:%M:%S')}")
        self._log_append(f"{'='*50}")

        def worker():
            try:
                count = convert_file(
                    inp, out,
                    progress_cb=lambda v: self.after(0, self._set_progress, v),
                    log_cb=lambda m: self.after(0, self._log_append, m)
                )
                self.after(0, self._done, count, out)
            except Exception as e:
                self.after(0, self._error, str(e))

        threading.Thread(target=worker, daemon=True).start()

    def _resolve_writable_path(self, preferred_path):
        """
        Megpróbálja megnyitni a kívánt mentési helyet írásra.
        Ha nem sikerül (pl. hálózati meghajtó, védett mappa),
        visszaesik: 1) forrás fájl mappája, 2) Asztal, 3) Letöltések, 4) Temp.
        """
        candidates = [preferred_path]

        # 1. forrás fájl mappájában
        inp = self._input_path.get().strip()
        if inp:
            src_dir = os.path.dirname(inp)
            fname = os.path.basename(preferred_path)
            candidates.append(os.path.join(src_dir, fname))

        # 2. Asztal
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        candidates.append(os.path.join(desktop, os.path.basename(preferred_path)))

        # 3. Letöltések
        downloads = os.path.join(os.path.expanduser("~"), "Downloads")
        candidates.append(os.path.join(downloads, os.path.basename(preferred_path)))

        # 4. Temp
        import tempfile
        candidates.append(os.path.join(tempfile.gettempdir(), os.path.basename(preferred_path)))

        for path in candidates:
            try:
                folder = os.path.dirname(path)
                if folder and not os.path.exists(folder):
                    os.makedirs(folder, exist_ok=True)
                # Teszt: megnyitható-e írásra?
                with open(path, 'ab') as f:
                    pass
                if path != preferred_path:
                    messagebox.showwarning(
                        "Mentési hely módosítva",
                        f"Az eredeti mentési hely nem elérhető (jogosultság vagy hálózati hiba).\n\n"
                        f"Mentés ide:\n{path}"
                    )
                return path
            except (PermissionError, OSError):
                continue

        # Ha semmi sem működik: visszaadjuk az eredetit (a hiba majd megjelenik)
        return preferred_path

    def _done(self, count, out_path):
        self._btn.configure(state='normal', text="▶  BELSŐ VERZIÓ GENERÁLÁSA")
        messagebox.showinfo("Kész!",
            f"Sikeresen létrehozva {count} munkalap.\n\nMentve:\n{out_path}")
        # Automatikusan megnyitja a fájlt
        try:
            os.startfile(out_path)
        except:
            pass

    def _error(self, msg):
        self._btn.configure(state='normal', text="▶  BELSŐ VERZIÓ GENERÁLÁSA")
        self._log_append(f"\n❌ HIBA: {msg}")
        messagebox.showerror("Hiba", f"Konverzió sikertelen:\n\n{msg}")


if __name__ == '__main__':
    app = App()
    app.mainloop()
